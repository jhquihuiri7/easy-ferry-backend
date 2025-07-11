from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, PatternFill
from reports.content_section import content_section
from reports.footer_section import footer_section
from openpyxl.styles.borders import Border, Side
from authentication.models import Crew, Owner
from reports.models import Business

def apply_borders(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

def generate_daily_report(sales, date, time):
    buffer = BytesIO()

    # Crear archivo y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Establecer orientación de página
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Establecer márgenes de página
    ws.page_margins = PageMargins(
        left=0.3,
        right=0.1,
        top=0.4,
        bottom=0,
        header=0,
        footer=0
    )

    # Establecer anchos de columna
    ws.column_dimensions['A'].width = 4.83
    ws.column_dimensions['B'].width = 4.83
    ws.column_dimensions['C'].width = 4.83
    ws.column_dimensions['D'].width = 10.35
    ws.column_dimensions['H'].width = 2.76
    ws.column_dimensions['M'].width = 3.45
    ws.column_dimensions['N'].width = 5.52
    ws.column_dimensions['T'].width = 4.83
    ws.column_dimensions['U'].width = 3.45
    ws.column_dimensions['W'].width = 4.5
    ws.column_dimensions['S'].width = 7.5

    # Rango A-Z a 4.5 (sobrescribe las anteriores A-C si lo haces después)
    for col in range(ord('A'), ord('Z') + 1):
        col_letter = chr(col)
        if col_letter not in ['A', 'B', 'C', 'D', 'H', 'M', 'N', 'T', 'U', 'W','S']:
            ws.column_dimensions[col_letter].width = 6.21

    # Establecer alturas de fila
    for i in range(1, 66):
        ws.row_dimensions[i].height = 12
    ws.row_dimensions[2].height = 24

    information_section(ws, date, time)
    content_section(ws, sales)
    footer_section(ws, 11+len(sales))

    apply_borders(ws)
    # Guardar en el buffer
    wb.save(buffer)
    buffer.seek(0)

    return buffer

def set_style(ws, cell_range, font=None, alignment=None, fill=None):
    for row in ws[cell_range]:
        for cell in row:
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill

def apply_font_style_subtitle(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name="Arial", size=9, bold=True)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center")

def apply_font_style_text(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name="Arial", size=8)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center")

def set_cell_style_subtitle(ws, start_cell, end_cell):
    fill = PatternFill(fill_type="solid", start_color="B8CCE4", end_color="B8CCE4")  # Azul claro, puedes cambiar el color
    ws.merge_cells(f"{start_cell}:{end_cell}")
    set_style(ws, f"{start_cell}:{end_cell}",
              font=Font(name="Arial", size=9, bold=True),
              alignment=Alignment(horizontal="left", vertical="center"),
              fill=fill)


def set_cell_style_text(ws, start_cell, end_cell):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    set_style(ws, f"{start_cell}:{end_cell}",
              font=Font(name="Arial", size=8),
              alignment=Alignment(horizontal="left", vertical="center"))

def information_section(ws, date, time_period):

    # Get the crew instance
    business_instance = Business.objects.get(business="Gaviota")
    crew = Crew.objects.get(business=business_instance.id)
    owner = Owner.objects.get(business=business_instance.id)

    if business_instance.business == "Gaviota" or business_instance.business == "Viamar":
        if time_period == "am":
            p_zarpe = "Pto. Baq. Moreno"
            p_arribo = "Pto. Ayora"
            h_zarpe = "07:00"
            h_arribo = "09:00"
        else:
            p_zarpe = "Pto. Ayora"
            p_arribo = "Pto. Baq. Moreno"
            h_zarpe = "15:00"
            h_arribo = "17:00"

    # Línea 1
    apply_font_style_subtitle(ws, "A1", "I. INFORMACIÓN DEL VIAJE")
    apply_font_style_subtitle(ws, "P1", "II. INFORMACIÓN DE LA NAVE")
    set_cell_style_subtitle(ws, "A1", "O1")
    set_cell_style_subtitle(ws, "P1", "W1")

    # Línea 2
    apply_font_style_subtitle(ws, "A2", "Fecha:")
    apply_font_style_text(ws, "C2", date)
    apply_font_style_subtitle(ws, "E2", "Puerto zarpe:")
    apply_font_style_text(ws, "H2", p_zarpe)
    apply_font_style_subtitle(ws, "K2", "Hora zarpe:")
    apply_font_style_text(ws, "N2", h_zarpe)
    set_cell_style_text(ws, "A2", "B2")
    set_cell_style_text(ws, "C2", "D2")
    set_cell_style_text(ws, "E2", "G2")
    set_cell_style_text(ws, "H2", "J2")
    set_cell_style_text(ws, "K2", "M2")
    set_cell_style_text(ws, "N2", "O2")

    apply_font_style_subtitle(ws, "A3", "Actividad:")
    apply_font_style_text(ws, "C3", "Cabotaje")
    apply_font_style_subtitle(ws, "E3", "Puerto arribo:")
    apply_font_style_text(ws, "H3", p_arribo)
    apply_font_style_subtitle(ws, "K3", "Hora arribo:")
    apply_font_style_text(ws, "N3", h_arribo)
    set_cell_style_text(ws, "A3", "B3")
    set_cell_style_text(ws, "C3", "D3")
    set_cell_style_text(ws, "E3", "G3")
    set_cell_style_text(ws, "H3", "J3")
    set_cell_style_text(ws, "K3", "M3")
    set_cell_style_text(ws, "N3", "O3")

    apply_font_style_subtitle(ws, "P2", "Nombre:")
    apply_font_style_text(ws, "R2", business_instance.ferry)
    apply_font_style_subtitle(ws, "T2", "Cap. Tripulantes:")
    apply_font_style_text(ws, "W2", crew.crew_capacity)
    set_cell_style_text(ws, "P2", "Q2")
    set_cell_style_text(ws, "R2", "S2")
    set_cell_style_text(ws, "T2", "V2")
    set_cell_style_text(ws, "W2", "W2")

    apply_font_style_subtitle(ws, "P3", "Matrícula:")
    apply_font_style_text(ws, "R3", crew.ferry_registration)
    apply_font_style_subtitle(ws, "T3", "Cap. Pasajeros:")
    apply_font_style_text(ws, "W3", crew.passenger_capacity)
    set_cell_style_text(ws, "P3", "Q3")
    set_cell_style_text(ws, "R3", "S3")
    set_cell_style_text(ws, "T3", "V3")
    set_cell_style_text(ws, "W3", "W3")

    # Línea 4
    apply_font_style_subtitle(ws, "A4", "III. INFORMACIÓN ARMADOR")
    apply_font_style_subtitle(ws, "H4", "IV. INFORMACIÓN RESPONSABLE DEL EMBARQUE")
    apply_font_style_subtitle(ws, "P4", "V. INFORMACIÓN TRIPULANTES")
    set_cell_style_subtitle(ws, "A4", "G4")
    set_cell_style_subtitle(ws, "H4", "O4")
    set_cell_style_subtitle(ws, "P4", "W4")

    # Línea 5
    apply_font_style_subtitle(ws, "A5", "Nombre:")
    apply_font_style_text(ws, "C5", owner.name)
    apply_font_style_subtitle(ws, "H5", "Nombre:")
    apply_font_style_text(ws, "J5", crew.responsible_name)
    apply_font_style_subtitle(ws, "P5", "Capitán:")
    apply_font_style_text(ws, "R5", crew.captain_name)
    apply_font_style_subtitle(ws, "T5", "Cédula:")
    apply_font_style_text(ws, "V5", crew.captain_passport)
    set_cell_style_text(ws, "A5", "B5")
    set_cell_style_text(ws, "C5", "G5")
    set_cell_style_text(ws, "H5", "I5")
    set_cell_style_text(ws, "J5", "O5")
    set_cell_style_text(ws, "P5", "Q5")
    set_cell_style_text(ws, "R5", "S5")
    set_cell_style_text(ws, "T5", "U5")
    set_cell_style_text(ws, "V5", "W5")

    # Línea 6
    apply_font_style_subtitle(ws, "A6", "RUC")
    apply_font_style_text(ws, "C6", owner.ruc)
    apply_font_style_subtitle(ws, "E6", "Teléf:")
    apply_font_style_text(ws, "F6", owner.phone)
    apply_font_style_subtitle(ws, "H6", "Cédula:")
    apply_font_style_text(ws, "J6", crew.responsible_passport)
    apply_font_style_subtitle(ws, "L6", "Teléfono:")
    apply_font_style_text(ws, "N6", crew.responsible_phone)
    apply_font_style_subtitle(ws, "P6", "Marinero 1:")
    apply_font_style_text(ws, "R6", crew.sailor1_name)
    apply_font_style_subtitle(ws, "T6", "Cédula:")
    apply_font_style_text(ws, "V6", crew.sailor1_passport)
    set_cell_style_text(ws, "A6", "B6")
    set_cell_style_text(ws, "C6", "D6")
    set_cell_style_text(ws, "E6", "E6")
    set_cell_style_text(ws, "F6", "G6")
    set_cell_style_text(ws, "H6", "I6")
    set_cell_style_text(ws, "J6", "K6")
    set_cell_style_text(ws, "L6", "M6")
    set_cell_style_text(ws, "N6", "O6")
    set_cell_style_text(ws, "P6", "Q6")
    set_cell_style_text(ws, "R6", "S6")
    set_cell_style_text(ws, "T6", "U6")
    set_cell_style_text(ws, "V6", "W6")

    # Línea 7
    apply_font_style_subtitle(ws, "A7", "e-mail:")
    apply_font_style_text(ws, "C7", owner.email)
    apply_font_style_subtitle(ws, "H7", "e-mail:")
    apply_font_style_text(ws, "J7", crew.responsible_email)
    apply_font_style_subtitle(ws, "P7", "")
    apply_font_style_text(ws, "R7", "")
    apply_font_style_subtitle(ws, "T7", "")
    apply_font_style_text(ws, "V7", "")
    set_cell_style_text(ws, "A7", "B7")
    set_cell_style_text(ws, "C7", "G7")
    set_cell_style_text(ws, "H7", "I7")
    set_cell_style_text(ws, "J7", "O7")
    set_cell_style_text(ws, "P7", "Q7")
    set_cell_style_text(ws, "R7", "S7")
    set_cell_style_text(ws, "T7", "U7")
    set_cell_style_text(ws, "V7", "W7")
