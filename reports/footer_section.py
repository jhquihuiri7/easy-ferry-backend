from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

def footer_section(ws: Worksheet, cell_start: int):
    # Alturas de fila
    ws.row_dimensions[cell_start].height = 22
    for i in range(cell_start + 1, cell_start + 8):
        ws.row_dimensions[i].height = 15

    # Línea 1 - Títulos de secciones
    apply_font_style_subtitle(ws, f"A{cell_start}", "VII. RESPONSABLE DEL REGISTRO DE PASAJEROS")
    apply_font_style_subtitle(ws, f"L{cell_start}", "VIII. CAPITANÍA DEL PUERTO (Zarpe)")
    apply_font_style_subtitle(ws, f"P{cell_start}", "IX. GAD MUNICIPAL (Recepción)")
    apply_font_style_subtitle(ws, f"T{cell_start}", "X. CAPITANIA DE PUERTO (Control)")

    set_cell_style_subtitle(ws, f"A{cell_start}", f"K{cell_start}")
    set_cell_style_subtitle(ws, f"L{cell_start}", f"O{cell_start}")
    set_cell_style_subtitle(ws, f"P{cell_start}", f"S{cell_start}")
    set_cell_style_subtitle(ws, f"T{cell_start}", f"W{cell_start}")

    # Línea 2 - Declaración y campos nombre
    apply_font_style_text(ws, f"A{cell_start + 1}", "Declaración de responsabilidad: El Armador asume toda responsabilidad legal sobre los actos relacionados con la operación de la embarcación, incluido el registro de pasajeros. Asimismo, como persona responsable del registro de pasajeros DECLARO que la información detallada en el presente formulario es verídica en su totalidad, asimismo, conozco que puede estar sujeto al análisis que en derecho corresponda y que es de mi entera responsabilidad cualquier tipo de falsificación, destrucción, adulteración, modificación u omisión en la información proporcionada a las Autoridades competentes.")
    ws[f"A{cell_start + 1}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{cell_start + 1}:G{cell_start + 7}")

    for col in ['H', 'L', 'P', 'T']:
        apply_font_style_subtitle(ws, f"{col}{cell_start + 1}", "Nombre:")
        set_cell_style_text(ws, f"{col}{cell_start + 1}", f"{chr(ord(col)+1)}{cell_start + 1}")  # Merge with next col

    # Línea 3 - Cédula
    for col in ['H', 'L', 'P', 'T']:
        apply_font_style_subtitle(ws, f"{col}{cell_start + 2}", "Cédula:")
        set_cell_style_text(ws, f"{col}{cell_start + 2}", f"{chr(ord(col)+1)}{cell_start + 2}")

    # Línea 4 - Cargo
    for col in ['H', 'L', 'P', 'T']:
        apply_font_style_subtitle(ws, f"{col}{cell_start + 3}", "Cargo:")
        set_cell_style_text(ws, f"{col}{cell_start + 3}", f"{chr(ord(col)+1)}{cell_start + 3}")

    # Línea 5 - Fecha
    for col in ['H', 'L', 'P', 'T']:
        apply_font_style_subtitle(ws, f"{col}{cell_start + 4}", "Fecha:")
        set_cell_style_text(ws, f"{col}{cell_start + 4}", f"{chr(ord(col)+1)}{cell_start + 4}")

    # Línea 6 - Teléfono
    for col in ['H', 'L', 'P', 'T']:
        apply_font_style_subtitle(ws, f"{col}{cell_start + 5}", "Teléfono:")
        set_cell_style_text(ws, f"{col}{cell_start + 5}", f"{chr(ord(col)+1)}{cell_start + 5}")

    # Línea 7 y 8 - Firma y sello (merge hacia abajo)
    for col in ['H', 'L', 'P', 'T']:
        apply_font_style_subtitle(ws, f"{col}{cell_start + 6}", "Firma y sello:")
        ws.merge_cells(f"{col}{cell_start + 6}:{chr(ord(col)+1)}{cell_start + 7}")
        set_cell_style_text(ws, f"{col}{cell_start + 6}", f"{chr(ord(col)+1)}{cell_start + 7}")


from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

def is_merged_cell(ws, cell_coordinate):
    for merged_range in ws.merged_cells.ranges:
        if cell_coordinate in merged_range:
            return True
    return False

def apply_font_style_subtitle(ws, cell, text):
    if not is_merged_cell(ws, cell):
        ws[cell] = text
    ws[cell].font = Font(name="Arial", size=9, bold=True)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_font_style_text(ws, cell, text):
    if not is_merged_cell(ws, cell):
        ws[cell] = text
    ws[cell].font = Font(name="Arial", size=8)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_font_style_mini(ws, cell, text):
    if not is_merged_cell(ws, cell):
        ws[cell] = text
    ws[cell].font = Font(name="Arial", size=8)
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

def set_style(ws, cell_range, font=None, alignment=None, fill=None):
    for row in ws[cell_range]:
        for cell in row:
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill

def set_cell_style_subtitle(ws, start_cell, end_cell, text=None):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    if text:
        ws[start_cell] = text
    fill = PatternFill(fill_type="solid", start_color="B8CCE4", end_color="B8CCE4")  # Azul claro, puedes cambiar el color
    set_style(
        ws,
        f"{start_cell}:{end_cell}",
        font=Font(name="Arial", size=8, bold=True),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        fill=fill
    )

def set_cell_style_text(ws, start_cell, end_cell, text=None):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    if text:
        ws[start_cell] = text
    set_style(
        ws,
        f"{start_cell}:{end_cell}",
        font=Font(name="Arial", size=8),
        alignment=Alignment(horizontal="left", vertical="center")
    )

def set_cell_style_mini(ws, start_cell, end_cell, text=None):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    if text:
        ws[start_cell] = text
    set_style(
        ws,
        f"{start_cell}:{end_cell}",
        font=Font(name="Arial", size=8),
        alignment=Alignment(wrap_text=True, vertical="top", horizontal="left")
    )

def apply_borders(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
