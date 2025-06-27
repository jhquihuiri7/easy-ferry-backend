from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def apply_font_style_subtitle(cell, value, ws):
    fill = PatternFill(fill_type="solid", start_color="B8CCE4", end_color="B8CCE4")  # Azul claro, puedes cambiar el color
    ws[cell] = value
    ws[cell].font = Font(name="Arial", size=9, bold=True)
    ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[cell].fill = fill

def apply_font_style_text(cell, value, ws):
    ws[cell] = value
    ws[cell].font = Font(name="Arial", size=8, bold=False)
    ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_merge_style(ws, start_cell, end_cell):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    # Optionally, set styles on merged cell
    ws[start_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[start_cell].font = Font(name="Arial", size=9, bold=True)

def content_section(ws, reserves):
    # Encabezado principal (fila 8)
    apply_font_style_subtitle("A8", "Check", ws)
    set_merge_style(ws, "A8", "B9")
    apply_font_style_subtitle("C8", "VI. REGISTRO DE PASAJEROS", ws)
    set_merge_style(ws, "C8", "W8")

    # Subencabezados (fila 9)
    apply_font_style_subtitle("A10", "Si", ws)
    apply_font_style_subtitle("B10", "No", ws)
    apply_font_style_subtitle("C9", "Nro", ws)
    set_merge_style(ws, "C9", "C10")
    apply_font_style_subtitle("D9", "Apellidos y Nombres (completos)", ws)
    set_merge_style(ws, "D9", "G10")
    apply_font_style_subtitle("H9", "Cédula/Pasaporte", ws)
    set_merge_style(ws, "H9", "J10")
    apply_font_style_subtitle("K9", "Nacionalidad", ws)
    set_merge_style(ws, "K9", "M10")
    apply_font_style_subtitle("N9", "Fecha de Nacimiento", ws)
    set_merge_style(ws, "N9", "O10")
    apply_font_style_subtitle("P9", "Estatus", ws)
    set_merge_style(ws, "P9", "R9")
    apply_font_style_subtitle("P10", "Res", ws)
    apply_font_style_subtitle("Q10", "Tra", ws)
    apply_font_style_subtitle("R10", "Tur", ws)
    apply_font_style_subtitle("S9", "Teléfono emergencia", ws)
    set_merge_style(ws, "S9", "T10")
    apply_font_style_subtitle("U9", "Observaciones", ws)
    set_merge_style(ws, "U9", "W10")

    # Cargar contenido a partir de la fila 11
    generate_content(ws, reserves)


def generate_content(ws, reserves):
    row_start = 11
    index = 0

    for i, reserve in enumerate(reserves):
        current_row = row_start + i
        age = max(reserve.age, 0)
        index_string = str(index + 1) if age >= 2 else ""

        if age >= 2:
            index += 1

        status = reserve.status
        res = "x" if status == "residente" else ""
        tem = "x" if status == "transeunte" else ""
        tur = "x" if status == "turista" else ""
        if not any([res, tem, tur]):
            res = "x"

        data = {
            f"C{current_row}": index_string,
            f"D{current_row}": reserve.name.upper(),
            f"H{current_row}": reserve.passport.upper(),
            f"K{current_row}": "Country".upper(),
            f"N{current_row}": str(age),
            f"P{current_row}": res,
            f"Q{current_row}": tem,
            f"R{current_row}": tur,
            f"S{current_row}": reserve.phone,
            f"U{current_row}": reserve.notes.upper(),
        }

        for cell, value in data.items():
            apply_font_style_text(cell, value, ws)

        # Merges por fila
        ws.merge_cells(f"A{current_row}:A{current_row}")
        ws.merge_cells(f"B{current_row}:B{current_row}")
        ws.merge_cells(f"C{current_row}:C{current_row}")
        ws.merge_cells(f"D{current_row}:G{current_row}")
        ws.merge_cells(f"H{current_row}:J{current_row}")
        ws.merge_cells(f"K{current_row}:M{current_row}")
        ws.merge_cells(f"N{current_row}:O{current_row}")
        ws.merge_cells(f"P{current_row}:P{current_row}")
        ws.merge_cells(f"Q{current_row}:Q{current_row}")
        ws.merge_cells(f"R{current_row}:R{current_row}")
        ws.merge_cells(f"S{current_row}:T{current_row}")
        ws.merge_cells(f"U{current_row}:W{current_row}")
